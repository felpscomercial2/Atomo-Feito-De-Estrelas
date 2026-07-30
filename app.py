from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import psycopg2
import psycopg2.extras
import os
import re
import time
import json
import io

app = Flask(__name__)
CORS(app)

# ============================================================
# CACHE SIMPLES EM MEMÓRIA
# Guarda resultados por 5 minutos para não bater no banco
# toda vez que alguém acessa a página
# ============================================================
_cache = {}
CACHE_TTL = 28800  # 8 horas em segundos

def cache_get(key):
    if key in _cache:
        valor, timestamp = _cache[key]
        if time.time() - timestamp < CACHE_TTL:
            return valor
    return None

def cache_set(key, valor):
    _cache[key] = (valor, time.time())

def cache_clear():
    _cache.clear()

# ============================================================
# CONEXÃO COM SUPABASE
# ============================================================
def get_conn():
    # Tenta conectar até 3 vezes — Supabase pode fechar conexões idle
    last_err = None
    for attempt in range(3):
        try:
            conn = psycopg2.connect(
                host            = os.environ.get('DB_HOST'),
                port            = int(os.environ.get('DB_PORT', 5432)),
                database        = os.environ.get('DB_NAME', 'railway'),
                user            = os.environ.get('DB_USER'),
                password        = os.environ.get('DB_PASS'),
                sslmode         = 'require',
                connect_timeout = 10,
                keepalives      = 1,
                keepalives_idle    = 30,
                keepalives_interval = 10,
                keepalives_count   = 5,
            )
            return conn
        except Exception as e:
            last_err = e
            time.sleep(0.5 * (attempt + 1))
    raise last_err

def _serializar_row(row):
    import datetime
    out = {}
    for k, v in row.items():
        if isinstance(v, datetime.date):
            try:
                out[k] = v.isoformat()
            except Exception:
                out[k] = None  # data inválida no banco (ex: ano 20256) → ignora
        else:
            out[k] = v
    return out

def consultar(sql, params=()):
    conn   = get_conn()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cursor.execute(sql, params)
    # Lê linha a linha para não quebrar em datas inválidas
    rows = []
    while True:
        try:
            row = cursor.fetchone()
            if row is None:
                break
            rows.append(_serializar_row(dict(row)))
        except Exception:
            continue  # pula linha com data inválida e segue
    cursor.close()
    conn.close()
    return rows

# ============================================================
# MONTA FILTROS
# ============================================================
def montar_filtros(args):
    condicoes = []
    params    = []

    anos = args.getlist('ano')
    if anos:
        placeholders = ','.join(['%s'] * len(anos))
        condicoes.append(f"ano IN ({placeholders})")
        params.extend([int(a) for a in anos])

    meses = args.getlist('mes')
    if meses:
        placeholders = ','.join(['%s'] * len(meses))
        condicoes.append(f"mes IN ({placeholders})")
        params.extend([int(m) for m in meses])

    unidade = args.get('unidade')
    if unidade:
        condicoes.append("unidade = %s")
        params.append(unidade)

    uf = args.get('uf')
    if uf:
        condicoes.append("uf = %s")
        params.append(uf)

    tipo = args.get('tipo')
    if tipo:
        condicoes.append("tipo_operacao = %s")
        params.append(tipo)

    marca = args.get('marca')
    if marca:
        condicoes.append("marca = %s")
        params.append(marca)

    vendedores = args.getlist('vendedor')
    if vendedores:
        # -------------------------------------------------------------
        # CORRECAO: filtrar pelo CODIGO do vendedor, nao pelo nome.
        # O mesmo vendedor pode aparecer com variacoes de nome
        # (espacos extras, maiuscula/minuscula, nome diferente por
        # unidade). Filtrar so pelo nome exato descartava notas e o
        # total ficava menor que a planilha.
        # Aqui pegamos todos os cod_vendedor ligados aos nomes
        # selecionados e trazemos TODAS as notas desses codigos.
        # -------------------------------------------------------------
        placeholders = ','.join(['%s'] * len(vendedores))
        nomes_norm   = [' '.join(str(v).strip().upper().split()) for v in vendedores]
        ph_norm      = ','.join(['%s'] * len(nomes_norm))
        condicoes.append(
            "("
            "  UPPER(BTRIM(vendedor)) IN (" + ph_norm + ")"
            "  OR (cod_vendedor IS NOT NULL AND cod_vendedor::TEXT <> '' AND cod_vendedor::TEXT IN ("
            "        SELECT DISTINCT f2.cod_vendedor::TEXT FROM faturamento f2"
            "         WHERE UPPER(BTRIM(f2.vendedor)) IN (" + ph_norm + ")"
            "           AND f2.cod_vendedor IS NOT NULL AND f2.cod_vendedor::TEXT <> ''"
            "        UNION"
            "        SELECT DISTINCT v2.cod_vendedor::TEXT FROM vendedores v2"
            "         WHERE UPPER(BTRIM(v2.nome)) IN (" + ph_norm + ")"
            "           AND v2.cod_vendedor IS NOT NULL AND v2.cod_vendedor::TEXT <> ''"
            "     ))"
            ")"
        )
        params.extend(nomes_norm)   # comparacao direta pelo nome normalizado
        params.extend(nomes_norm)   # subquery faturamento
        params.extend(nomes_norm)   # subquery vendedores

    where = ("WHERE " + " AND ".join(condicoes)) if condicoes else ""
    return where, params

def cache_key(rota, args):
    return rota + '?' + '&'.join(f'{k}={v}' for k, v in sorted(args.items()))

# ============================================================
# ROTAS
# ============================================================
def _aquecer_cache():
    import threading
    def _warm():
        try:
            conn   = get_conn()
            cursor = conn.cursor()
            def q(sql):
                cursor.execute(sql)
                return [r[0] for r in cursor.fetchall()]
            filtros = {
                'anos':       q("SELECT DISTINCT ano FROM faturamento WHERE ano IS NOT NULL AND ano > 0 ORDER BY ano DESC"),
                'meses':      q("SELECT DISTINCT mes FROM faturamento WHERE mes IS NOT NULL AND mes > 0 ORDER BY mes"),
                'unidades':   q("SELECT DISTINCT unidade FROM faturamento WHERE unidade IS NOT NULL ORDER BY unidade"),
                'ufs':        q("SELECT DISTINCT uf FROM faturamento WHERE uf IS NOT NULL AND uf != '' ORDER BY uf"),
                'marcas':     q("SELECT DISTINCT marca FROM faturamento WHERE marca IS NOT NULL ORDER BY marca"),
                'tipos':      q("SELECT DISTINCT tipo_operacao FROM faturamento WHERE tipo_operacao IS NOT NULL ORDER BY tipo_operacao"),
                'vendedores': q("SELECT DISTINCT vendedor FROM faturamento WHERE vendedor IS NOT NULL ORDER BY vendedor"),
            }
            cache_set('filtros', filtros)
            cursor2 = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cursor2.execute("SELECT DISTINCT produto, cod_produto, marca FROM faturamento WHERE produto IS NOT NULL AND produto != '' ORDER BY produto")
            produtos = [_serializar_row(dict(r)) for r in cursor2.fetchall()]
            cache_set('todos_produtos', produtos)
            cursor.close(); cursor2.close(); conn.close()
            print("Cache aquecido com sucesso!")
        except Exception as e:
            print(f"Erro cache warmup: {e}")
    threading.Thread(target=_warm, daemon=True).start()

# Aquece cache ao iniciar o servidor
_aquecer_cache()

@app.route('/')
def home():
    return jsonify({"status": "online", "mensagem": "API Átomo funcionando!"})

@app.route('/api/filtros')
def filtros():
    key    = 'filtros'
    cached = cache_get(key)
    if cached: return jsonify(cached)

    # Uma conexão, queries executadas sequencialmente
    conn   = get_conn()
    cursor = conn.cursor()

    def q(sql):
        cursor.execute(sql)
        return [r[0] for r in cursor.fetchall()]

    resultado = {
        'anos':       q("SELECT DISTINCT ano  FROM faturamento WHERE ano  IS NOT NULL AND ano > 0 ORDER BY ano DESC"),
        'meses':      q("SELECT DISTINCT mes  FROM faturamento WHERE mes  IS NOT NULL AND mes > 0 ORDER BY mes"),
        'unidades':   q("SELECT DISTINCT unidade FROM faturamento WHERE unidade IS NOT NULL ORDER BY unidade"),
        'ufs':        q("SELECT DISTINCT uf   FROM faturamento WHERE uf   IS NOT NULL AND uf != '' ORDER BY uf"),
        'marcas':     q("SELECT DISTINCT marca FROM faturamento WHERE marca IS NOT NULL ORDER BY marca"),
        'tipos':      q("SELECT DISTINCT tipo_operacao FROM faturamento WHERE tipo_operacao IS NOT NULL ORDER BY tipo_operacao"),
        'vendedores': q("SELECT DISTINCT vendedor FROM faturamento WHERE vendedor IS NOT NULL ORDER BY vendedor"),
    }
    cursor.close(); conn.close()
    cache_set(key, resultado)
    return jsonify(resultado)


@app.route('/api/dashboard')
def dashboard():
    """Retorna KPIs + mensal + unidade + vendedores + marcas em UMA chamada."""
    key    = cache_key('dashboard', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    and_or = 'AND' if where else 'WHERE'

    conn   = get_conn()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    def run(sql, p):
        cursor.execute(sql, p)
        return [_serializar_row(dict(r)) for r in cursor.fetchall()]

    kpis = run(f"""
        SELECT
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Venda' THEN valor_nf ELSE 0 END) AS NUMERIC),2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END) AS NUMERIC),2) AS devolucoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Bonificacao' THEN valor_nf ELSE 0 END) AS NUMERIC),2) AS bonificacoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Venda' THEN valor_nf ELSE 0 END)/NULLIF(COUNT(CASE WHEN tipo_operacao='Venda' THEN 1 END),0) AS NUMERIC),2) AS ticket_medio,
            COUNT(DISTINCT cliente) AS total_clientes,
            COUNT(CASE WHEN tipo_operacao='Venda' THEN 1 END) AS qtd_vendas
        FROM faturamento {where}
    """, params)

    mensal = run(f"""
        SELECT ano, mes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Venda' THEN valor_nf ELSE 0 END) AS NUMERIC),2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END) AS NUMERIC),2) AS devolucoes
        FROM faturamento {where} {and_or} mes > 0
        GROUP BY ano, mes ORDER BY ano, mes
    """, params)

    unidade = run(f"""
        SELECT unidade,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Venda' THEN valor_nf ELSE 0 END) AS NUMERIC),2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END) AS NUMERIC),2) AS devolucoes,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} unidade IS NOT NULL
        GROUP BY unidade ORDER BY faturamento DESC
    """, params)

    vendedores = run(f"""
        SELECT vendedor,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC),2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao='Venda'
        GROUP BY vendedor ORDER BY faturamento DESC LIMIT 10
    """, params)

    marcas = run(f"""
        SELECT marca,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC),2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao='Venda' AND marca IS NOT NULL
        GROUP BY marca ORDER BY faturamento DESC LIMIT 15
    """, params)

    cursor.close(); conn.close()

    resultado = {
        'kpis':       kpis[0] if kpis else {},
        'mensal':     mensal,
        'unidade':    unidade,
        'vendedores': vendedores,
        'marcas':     marcas,
    }
    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/kpis')
def kpis():
    key    = cache_key('kpis', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    resultado = consultar(f"""
        SELECT
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Devolucao' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS devolucoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Bonificacao' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS bonificacoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) /
                NULLIF(COUNT(CASE WHEN tipo_operacao = 'Venda' THEN 1 END), 0) AS NUMERIC), 2) AS ticket_medio,
            COUNT(DISTINCT cliente) AS total_clientes,
            COUNT(CASE WHEN tipo_operacao = 'Venda' THEN 1 END) AS qtd_vendas
        FROM faturamento {where}
    """, params)

    r = resultado[0] if resultado else {}
    cache_set(key, r)
    return jsonify(r)

@app.route('/api/faturamento-mensal')
def faturamento_mensal():
    key    = cache_key('faturamento-mensal', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    resultado = consultar(f"""
        SELECT ano, mes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Devolucao' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS devolucoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Bonificacao' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS bonificacoes
        FROM faturamento {where}
        {'AND' if where else 'WHERE'} mes > 0
        GROUP BY ano, mes ORDER BY ano, mes
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/top-vendedores')
def top_vendedores():
    key    = cache_key('top-vendedores', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 10))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT vendedor,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes,
            COUNT(DISTINCT unidade) AS unidades,
            COUNT(*) AS qtd_vendas
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY vendedor ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-marca')
def faturamento_por_marca():
    key    = cache_key('faturamento-por-marca', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 15))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT marca,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY marca ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-regiao')
def faturamento_por_regiao():
    key    = cache_key('faturamento-por-regiao', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT regiao,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        AND regiao IS NOT NULL AND regiao != ''
        GROUP BY regiao ORDER BY faturamento DESC
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-unidade')
def faturamento_por_unidade():
    key    = cache_key('faturamento-por-unidade', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT unidade,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Venda' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Devolucao' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS devolucoes,
            ROUND(CAST(SUM(CASE WHEN tipo_operacao = 'Bonificacao' THEN valor_nf ELSE 0 END) AS NUMERIC), 2) AS bonificacoes,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} unidade IS NOT NULL
        GROUP BY unidade ORDER BY faturamento DESC
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/top-produtos')
def top_produtos():
    key    = cache_key('top-produtos', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 10))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT produto, marca,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            ROUND(CAST(SUM(quantidade) AS NUMERIC), 0) AS quantidade
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY produto, marca ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-uf')
def faturamento_por_uf():
    key    = cache_key('faturamento-por-uf', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT uf,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        AND uf IS NOT NULL AND uf != ''
        GROUP BY uf ORDER BY faturamento DESC
    """, params)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/carteira-vendedor')
def carteira_vendedor():
    key    = cache_key('carteira-vendedor', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    cod_vendedor = request.args.get('cod_vendedor', '')
    if cod_vendedor:
        resultado = consultar("""
            SELECT * FROM carteira WHERE cod_vendedor = %s ORDER BY cliente
        """, [cod_vendedor])
    else:
        resultado = consultar("""
            SELECT cod_vendedor, COUNT(*) as total_clientes
            FROM carteira GROUP BY cod_vendedor ORDER BY total_clientes DESC
        """)

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/resumo-carteira')
def resumo_carteira():
    """
    Retorna total de clientes em carteira e margem média.
    Se filtrar por vendedor (nome), busca o cod_vendedor correspondente
    e retorna a carteira daquele vendedor.
    """
    key    = cache_key('resumo-carteira', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    vendedor     = request.args.get('vendedor', '')
    where, params = montar_filtros(request.args)
    and_or        = 'AND' if where else 'WHERE'

    margem = consultar(f"""
        SELECT ROUND(CAST(AVG(margem) AS NUMERIC), 2) AS margem_media
        FROM faturamento {where}
        {and_or} tipo_operacao = 'Venda'
        AND margem IS NOT NULL AND margem != 0
    """, params)
    margem_media = margem[0]['margem_media'] if margem else 0

    # Conta CNPJs distintos (mesmo CNPJ via Esdel e Indústria = 1 cliente único)
    if vendedor:
        cods = consultar("""
            SELECT DISTINCT cod_vendedor FROM faturamento
            WHERE vendedor = %s AND cod_vendedor IS NOT NULL AND cod_vendedor != ''
        """, [vendedor])
        if cods:
            cod_list     = [c['cod_vendedor'] for c in cods]
            placeholders = ','.join(['%s'] * len(cod_list))
            carteira     = consultar(f"""
                SELECT
                    COUNT(*)                                            AS total_codigos,
                    COUNT(DISTINCT NULLIF(TRIM(cnpj_cpf), ''))         AS total_cnpjs
                FROM carteira
                WHERE cod_vendedor IN ({placeholders})
            """, cod_list)
        else:
            carteira = [{'total_codigos': 0, 'total_cnpjs': 0}]
    else:
        carteira = consultar("""
            SELECT
                COUNT(*)                                            AS total_codigos,
                COUNT(DISTINCT NULLIF(TRIM(cnpj_cpf), ''))         AS total_cnpjs
            FROM carteira
        """)

    total_codigos  = carteira[0]['total_codigos'] if carteira else 0
    total_cnpjs    = carteira[0]['total_cnpjs']   if carteira else 0
    resultado = {
        'total_carteira': total_cnpjs,    # CNPJs únicos — usado no KPI principal
        'total_codigos':  total_codigos,  # todos os códigos (incluindo duplicatas por unidade)
        'margem_media':   float(margem_media) if margem_media else 0,
    }
    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/top-clientes')
def top_clientes():
    key    = cache_key('top-clientes', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 10))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT cliente,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(*) AS qtd_vendas
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        GROUP BY cliente ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/faturamento-por-cidade')
def faturamento_por_cidade():
    key    = cache_key('faturamento-por-cidade', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    limite   = int(request.args.get('limite', 15))
    and_or   = 'AND' if where else 'WHERE'
    resultado = consultar(f"""
        SELECT cidade, uf,
            ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
            COUNT(DISTINCT cliente) AS clientes
        FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
        AND cidade IS NOT NULL AND cidade != ''
        GROUP BY cidade, uf ORDER BY faturamento DESC LIMIT %s
    """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/todos-produtos')
def todos_produtos():
    key    = 'todos_produtos'
    cached = cache_get(key)
    if cached: return jsonify(cached)

    resultado = consultar("""
        SELECT DISTINCT produto, cod_produto, marca
        FROM faturamento
        WHERE produto IS NOT NULL AND produto != ''
        ORDER BY produto
    """)
    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/buscar-produtos')
def buscar_produtos():
    termo = request.args.get('q', '').strip()
    if not termo or len(termo) < 2:
        return jsonify([])

    # Cache por termo para evitar queries repetidas
    cache_k = f'busca_prod_{termo.lower()}'
    cached  = cache_get(cache_k)
    if cached: return jsonify(cached)

    # Busca por prefixo primeiro (mais rápido), depois por substring
    resultado = consultar("""
        SELECT produto, cod_produto, marca FROM (
            SELECT DISTINCT produto, cod_produto, marca
            FROM faturamento
            WHERE produto IS NOT NULL AND produto != ''
            AND (
                LOWER(produto) LIKE LOWER(%s)
                OR LOWER(cod_produto) LIKE LOWER(%s)
            )
        ) sub
        ORDER BY
            CASE WHEN LOWER(produto) LIKE LOWER(%s) THEN 0 ELSE 1 END,
            produto
        LIMIT 20
    """, [f'%{termo}%', f'{termo}%', f'{termo}%'])

    cache_set(cache_k, resultado)
    return jsonify(resultado)

@app.route('/api/top-produtos-filtrado')
def top_produtos_filtrado():
    key    = cache_key('top-produtos-filtrado', dict(request.args))
    cached = cache_get(key)
    if cached: return jsonify(cached)

    where, params = montar_filtros(request.args)
    produtos = request.args.getlist('produtos')
    and_or   = 'AND' if where else 'WHERE'

    if produtos:
        placeholders = ','.join(['%s'] * len(produtos))
        resultado = consultar(f"""
            SELECT produto, marca,
                ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
                ROUND(CAST(SUM(quantidade) AS NUMERIC), 0) AS quantidade,
                COUNT(DISTINCT cliente) AS clientes
            FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
            AND produto IN ({placeholders})
            GROUP BY produto, marca ORDER BY faturamento DESC
        """, params + produtos)
    else:
        limite    = int(request.args.get('limite', 20))
        resultado = consultar(f"""
            SELECT produto, marca,
                ROUND(CAST(SUM(valor_nf) AS NUMERIC), 2) AS faturamento,
                ROUND(CAST(SUM(quantidade) AS NUMERIC), 0) AS quantidade,
                COUNT(DISTINCT cliente) AS clientes
            FROM faturamento {where} {and_or} tipo_operacao = 'Venda'
            GROUP BY produto, marca ORDER BY faturamento DESC LIMIT %s
        """, params + [limite])

    cache_set(key, resultado)
    return jsonify(resultado)

@app.route('/api/vendedores-por-produto')
def vendedores_por_produto():
    produtos = request.args.getlist('produtos')
    if not produtos:
        return jsonify([])
    key    = 'vend_prod_' + '_'.join(sorted(produtos))
    cached = cache_get(key)
    if cached: return jsonify(cached)
    placeholders = ','.join(['%s'] * len(produtos))
    resultado = consultar(f"""
        SELECT DISTINCT vendedor
        FROM faturamento
        WHERE produto IN ({placeholders})
        AND vendedor IS NOT NULL AND vendedor != ''
        AND tipo_operacao = 'Venda'
        ORDER BY vendedor
    """, produtos)
    res = [r['vendedor'] for r in resultado]
    cache_set(key, res)
    return jsonify(res)

# ============================================================
# SHELF LIFE — MIGRAÇÃO (rode uma vez após deploy)
# ============================================================
@app.route('/api/shelflife/migrar', methods=['GET', 'POST'])
def shelflife_migrar():
    """
    Cria a constraint UNIQUE necessária para o UPSERT funcionar.
    Chame este endpoint UMA VEZ após o deploy via:
      curl -X POST https://horus-bmcj.onrender.com/api/shelflife/migrar
    """
    conn   = get_conn()
    cursor = conn.cursor()
    resultados = []

    # 1. Remove duplicatas antes de criar a constraint (mantém o de maior id)
    cursor.execute("""
        DELETE FROM shelflife a
        USING shelflife b
        WHERE a.id < b.id
          AND a.semana      = b.semana
          AND a.unidade     = b.unidade
          AND a.cod_produto = b.cod_produto
          AND COALESCE(a.validade::text, '') = COALESCE(b.validade::text, '')
    """)
    removidos = cursor.rowcount
    resultados.append(f'Duplicatas removidas: {removidos}')

    # 2. Cria a constraint se não existir
    cursor.execute("""
        SELECT COUNT(*) FROM pg_constraint
        WHERE conname = 'shelflife_semana_unidade_cod_produto_validade_key'
    """)
    existe = cursor.fetchone()[0]

    if not existe:
        cursor.execute("""
            ALTER TABLE shelflife
            ADD CONSTRAINT shelflife_semana_unidade_cod_produto_validade_key
            UNIQUE (semana, unidade, cod_produto, validade)
        """)
        resultados.append('Constraint UNIQUE criada com sucesso')
    else:
        resultados.append('Constraint UNIQUE já existia')

    conn.commit(); cursor.close(); conn.close()
    return jsonify({'ok': True, 'detalhes': resultados})

# ============================================================
# SHELF LIFE — CONTROLE DE ACESSO
# ============================================================
# Lista de e-mails autorizados a acessar o Shelf Life
# Para adicionar ou remover alguém, edite esta lista
EMAILS_AUTORIZADOS_SL = [
    'comercial2@reforpan.com.br',
    'comercial3@esdel.com.br',
    'comercial1@esdel'
]

@app.route('/api/shelflife/verificar-acesso', methods=['POST'])
def shelflife_verificar_acesso():
    data  = request.get_json(force=True)
    email = str(data.get('email', '')).strip().lower()
    autorizado = email in [e.lower() for e in EMAILS_AUTORIZADOS_SL]
    return jsonify({'autorizado': autorizado, 'email': email})

# ============================================================
# SHELF LIFE
# ============================================================
from datetime import date as _date

@app.route('/api/shelflife/upload', methods=['POST'])
def shelflife_upload():
    data     = request.get_json()
    semana   = data.get('semana')
    unidade  = data.get('unidade')
    produtos = data.get('produtos', [])

    if not produtos:
        return jsonify({'erro': 'Nenhum produto enviado'}), 400

    conn   = get_conn()
    cursor = conn.cursor()

    hoje      = _date.today()
    inseridos = 0
    atualizados = 0

    for p in produtos:
        validade = p.get('validade', '')
        val_fmt  = validade[:10] if validade and len(str(validade)) >= 10 else None

        try:
            from datetime import datetime
            val_date = datetime.strptime(str(validade)[:10], '%Y-%m-%d').date()
            dias     = (val_date - hoje).days
        except:
            dias = 999

        if dias <= 30:   status = 'CRITICO'
        elif dias <= 60: status = 'ATENCAO'
        else:            status = 'OK'

        nome  = str(p.get('produto', ''))
        # SL somente quando cod_sl for um codigo numerico real (ignora '-', '', None)
        cod_sl_raw = str(p.get('cod_sl') or '').strip()
        is_sl = bool(cod_sl_raw and cod_sl_raw not in ('-', u'—', 'nan', 'none', 'null', '0'))

        # UPSERT: insere se não existe, atualiza dados logísticos se já existe.
        # Campos editoriais (acao, obs, vendedor, qtd_atual, vendas, data_inc)
        # são preservados via COALESCE — só atualizam se ainda estiverem NULL no banco.
        cursor.execute("""
            INSERT INTO shelflife (
                semana, unidade, cod_produto, cod_sl, produto, marca,
                quantidade_log, validade, dias_vencimento,
                vence_em, status_logistica, status, is_sl, valor_sl
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (semana, unidade, cod_produto, validade)
            DO UPDATE SET
                -- Dados logísticos: sempre atualiza com o novo arquivo
                cod_sl           = EXCLUDED.cod_sl,
                produto          = EXCLUDED.produto,
                marca            = EXCLUDED.marca,
                quantidade_log   = EXCLUDED.quantidade_log,
                dias_vencimento  = EXCLUDED.dias_vencimento,
                vence_em         = EXCLUDED.vence_em,
                status_logistica = EXCLUDED.status_logistica,
                status           = EXCLUDED.status,
                is_sl            = EXCLUDED.is_sl,
                updated_at       = NOW(),
                -- Campos editoriais: preserva o que já foi preenchido
                quantidade_atual    = COALESCE(shelflife.quantidade_atual,    EXCLUDED.quantidade_atual),
                venda_3meses        = COALESCE(shelflife.venda_3meses,        EXCLUDED.venda_3meses),
                venda_mes           = COALESCE(shelflife.venda_mes,           EXCLUDED.venda_mes),
                data_inconsistencia = COALESCE(shelflife.data_inconsistencia, EXCLUDED.data_inconsistencia),
                obs_logistica       = COALESCE(shelflife.obs_logistica,       EXCLUDED.obs_logistica),
                obs_gerais          = COALESCE(shelflife.obs_gerais,          EXCLUDED.obs_gerais),
                acao                = COALESCE(shelflife.acao,                EXCLUDED.acao),
                vendedor            = COALESCE(shelflife.vendedor,            EXCLUDED.vendedor),
                resolvido           = COALESCE(shelflife.resolvido,           EXCLUDED.resolvido),
                valor_sl            = COALESCE(shelflife.valor_sl,            EXCLUDED.valor_sl)
        """, [
            semana, unidade,
            p.get('cod_produto'), p.get('cod_sl'),
            nome, p.get('marca'),
            p.get('quantidade'), val_fmt,
            dias, p.get('vence_em'), p.get('status_logistica'),
            status, is_sl, None
        ])

        # xmax = 0 significa INSERT, > 0 significa UPDATE
        cursor.execute("SELECT xmax FROM shelflife WHERE semana=%s AND unidade=%s AND cod_produto=%s AND validade=%s",
                       [semana, unidade, p.get('cod_produto'), val_fmt])
        row = cursor.fetchone()
        if row and row[0] and int(row[0]) > 0:
            atualizados += 1
        else:
            inseridos += 1

    conn.commit(); cursor.close(); conn.close()
    return jsonify({
        'inseridos':   inseridos,
        'atualizados': atualizados,
        'total':       inseridos + atualizados,
        'semana':      semana,
        'unidade':     unidade
    })

@app.route('/api/shelflife/listar')
def shelflife_listar():
    import datetime as _dt, traceback as _tb
    try:
        semana  = request.args.get('semana')
        unidade = request.args.get('unidade')
        status  = request.args.get('status')
        where   = []; params = []

        if semana:
            where.append('semana = %s'); params.append(semana)
        else:
            where.append('semana = (SELECT MAX(semana) FROM shelflife)')

        if unidade:
            where.append('unidade = %s'); params.append(unidade)

        where_str = 'WHERE ' + ' AND '.join(where) if where else ''
        resultado = consultar('SELECT * FROM shelflife ' + where_str + ' ORDER BY validade ASC NULLS LAST', params)

        hoje = _dt.date.today()
        status_map = []
        for row in resultado:
            val = row.get('validade')
            if val:
                try:
                    val_date = _dt.date.fromisoformat(str(val)[:10])
                    dias = (val_date - hoje).days
                except Exception:
                    dias = row.get('dias_vencimento', 999)
            else:
                dias = row.get('dias_vencimento', 999)

            if dias <= 30:   novo_status = 'CRITICO'
            elif dias <= 60: novo_status = 'ATENCAO'
            else:            novo_status = 'OK'

            row['dias_vencimento'] = dias
            if not row.get('is_sl'):
                row['status'] = novo_status
            status_map.append(row)

        if status == 'SL':
            status_map = [r for r in status_map if r.get('is_sl')]
        elif status:
            status_map = [r for r in status_map if r.get('status') == status and not r.get('is_sl')]

        status_map.sort(key=lambda r: r.get('dias_vencimento', 9999))
        return jsonify(status_map)

    except Exception as e:
        return jsonify({'erro': str(e), 'trace': _tb.format_exc()}), 500

@app.route('/api/shelflife/migrar-valor-sl', methods=['POST'])
def shelflife_migrar_valor_sl():
    """Adiciona coluna valor_sl na tabela shelflife se não existir."""
    try:
        conn   = get_conn()
        cursor = conn.cursor()
        cursor.execute("""
            ALTER TABLE shelflife ADD COLUMN IF NOT EXISTS valor_sl NUMERIC(10,2)
        """)
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'ok': True, 'mensagem': 'Coluna valor_sl criada com sucesso'})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/shelflife/corrigir-is-sl', methods=['POST'])
def shelflife_corrigir_is_sl():
    """Corrige is_sl: True somente quando cod_sl for um codigo real (nao vazio, nao traco)."""
    try:
        conn   = get_conn()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE shelflife
            SET is_sl = CASE
                WHEN cod_sl IS NOT NULL
                 AND TRIM(cod_sl) <> ''
                 AND TRIM(cod_sl) NOT IN ('-', '—', 'nan', 'none', 'null', '0')
                THEN TRUE
                ELSE FALSE
            END
        """)
        atualizados = cursor.rowcount
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'ok': True, 'registros_corrigidos': atualizados})
    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/api/vendedores/sincronizar', methods=['POST'])
def vendedores_sincronizar():
    """Recebe lista [{cod_vendedor, nome}] e faz upsert na tabela vendedores."""
    try:
        data = request.get_json(force=True)
        lista = data.get('vendedores', [])
        if not lista:
            return jsonify({'erro': 'Lista vazia'}), 400
        conn = get_conn(); cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS vendedores (
                id           SERIAL PRIMARY KEY,
                cod_vendedor TEXT UNIQUE,
                nome         TEXT
            )
        """)
        for v in lista:
            cursor.execute("""
                INSERT INTO vendedores (cod_vendedor, nome)
                VALUES (%s, %s)
                ON CONFLICT (cod_vendedor) DO UPDATE SET nome = EXCLUDED.nome
            """, [str(v['cod_vendedor']), str(v['nome'])])
        conn.commit(); cursor.close(); conn.close()
        return jsonify({'ok': True, 'sincronizados': len(lista)})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/shelflife/semanas')
def shelflife_semanas():
    resultado = consultar("""
        SELECT DISTINCT semana, unidade, COUNT(*) as total
        FROM shelflife GROUP BY semana, unidade ORDER BY semana DESC
    """)
    return jsonify(resultado)

@app.route('/api/shelflife/atualizar', methods=['POST'])
def shelflife_atualizar():
    data    = request.get_json()
    id_prod = data.get('id')
    conn    = get_conn()
    cursor  = conn.cursor()

    cursor.execute("""
        SELECT semana, unidade, cod_produto, produto, quantidade_log,
               quantidade_atual, venda_3meses, venda_mes,
               data_inconsistencia, obs_logistica, obs_gerais, acao, vendedor
        FROM shelflife WHERE id = %s
    """, [id_prod])
    row = cursor.fetchone()

    # data_inconsistencia:
    #   None        → preserva o valor atual do banco (campo não foi tocado)
    #   '' (vazio)  → apaga intencionalmente (SET NULL)
    #   'yyyy-mm-dd'→ atualiza para a nova data
    _data_inc_raw = data.get('data_inconsistencia')
    if _data_inc_raw is None:
        _data_inc_sql    = 'data_inconsistencia'
        _data_inc_params = []
    elif _data_inc_raw == '':
        _data_inc_sql    = 'NULL'
        _data_inc_params = []
    else:
        # Valida a data antes de salvar — rejeita anos fora do intervalo 2000-2099
        try:
            import datetime as _dti
            _d = _dti.date.fromisoformat(str(_data_inc_raw)[:10])
            if not (2000 <= _d.year <= 2099):
                raise ValueError(f'Ano inválido: {_d.year}')
            _data_inc_raw = _d.isoformat()  # normaliza para yyyy-mm-dd
        except Exception:
            _data_inc_raw = None  # data inválida → não salva
        if _data_inc_raw:
            _data_inc_sql    = '%s::date'
            _data_inc_params = [_data_inc_raw]
        else:
            _data_inc_sql    = 'data_inconsistencia'  # mantém o que tem
            _data_inc_params = []

    cursor.execute(f"""
        UPDATE shelflife SET
            quantidade_atual     = %s,
            venda_3meses         = %s,
            venda_mes            = %s,
            data_inconsistencia  = {_data_inc_sql},
            obs_logistica        = %s,
            obs_gerais           = %s,
            acao                 = %s,
            vendedor             = %s,
            resolvido            = %s,
            valor_sl             = COALESCE(%s, valor_sl),
            cod_sl               = CASE WHEN %s IS NOT NULL THEN %s ELSE cod_sl END,
            produto              = CASE WHEN %s IS NOT NULL THEN %s ELSE produto END,
            is_sl                = CASE WHEN %s IS NOT NULL THEN %s ELSE is_sl END,
            updated_at           = NOW()
        WHERE id = %s
    """, [
        data.get('quantidade_atual'),
        data.get('venda_3meses'),
        data.get('venda_mes'),
        *_data_inc_params,
        data.get('obs_logistica'),
        data.get('obs_gerais'),
        data.get('acao'),
        data.get('vendedor'),
        data.get('resolvido', False),
        data.get('valor_sl'),
        data.get('cod_sl'),  data.get('cod_sl'),
        data.get('produto'), data.get('produto'),
        data.get('cod_sl'),  bool(data.get('cod_sl')),
        id_prod
    ])

    if row:
        campos = [
            ('quantidade_atual',    row[5],              data.get('quantidade_atual'),           'Qtde Atual'),
            ('venda_3meses',        row[6],              data.get('venda_3meses'),               'Venda 3 Meses'),
            ('venda_mes',           row[7],              data.get('venda_mes'),                  'Venda Mensal'),
            ('data_inconsistencia', str(row[8]) if row[8] else '', data.get('data_inconsistencia') or '', 'Data Inconsistencia'),
            ('obs_logistica',       row[9],              data.get('obs_logistica'),              'Obs. Logistica'),
            ('obs_gerais',          row[10],             data.get('obs_gerais'),                 'Obs. Gerais'),
            ('acao',                row[11],             data.get('acao'),                       'Acao'),
            ('vendedor',            row[12],             data.get('vendedor'),                   'Vendedor'),
        ]
        alteracoes = []
        for campo, antes, depois, label in campos:
            antes_str  = str(antes  or '').strip()
            depois_str = str(depois or '').strip()
            if antes_str != depois_str:
                alteracoes.append(f"{label}: [{antes_str or '—'}] → [{depois_str or '—'}]")

        if alteracoes:
            cursor.execute("""
                INSERT INTO shelflife_historico (
                    shelflife_id, semana, unidade, cod_produto, produto,
                    quantidade_log, quantidade_atual, venda_3meses, venda_mes,
                    acao, vendedor, obs_logistica, obs_gerais,
                    data_inconsistencia, usuario
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, [
                id_prod, row[0], row[1], row[2], row[3],
                row[4],
                data.get('quantidade_atual'),
                data.get('venda_3meses'),
                data.get('venda_mes'),
                data.get('acao'),
                data.get('vendedor'),
                '|'.join(alteracoes),
                data.get('obs_gerais'),
                data.get('data_inconsistencia') or None,
                data.get('usuario', 'admin')
            ])

    conn.commit(); cursor.close(); conn.close()
    return jsonify({'ok': True, 'alteracoes': alteracoes if row else []})

@app.route('/api/shelflife/excluir', methods=['POST'])
def shelflife_excluir():
    data    = request.get_json()
    semana  = data.get('semana')
    unidade = data.get('unidade')
    conn    = get_conn()
    cursor  = conn.cursor()

    if unidade:
        cursor.execute("DELETE FROM shelflife WHERE semana = %s AND unidade = %s", [semana, unidade])
    else:
        cursor.execute("DELETE FROM shelflife WHERE semana = %s", [semana])

    deleted = cursor.rowcount
    conn.commit(); cursor.close(); conn.close()
    return jsonify({'excluidos': deleted})

# ============================================================
# SHELF LIFE — EXPORTAR EXCEL FORMATADO
# ============================================================
@app.route('/api/shelflife/exportar', methods=['POST'])
def shelflife_exportar():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data   = request.get_json(force=True)
    linhas = data.get('linhas', [])
    if not linhas:
        return jsonify({'erro': 'Nenhum dado para exportar'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = 'Shelf Life'

    HEADER_BG = '1A1A2E'; HEADER_FG = 'FFFFFF'
    ROW_ODD   = 'F7F8FA'; ROW_EVEN  = 'FFFFFF'
    STATUS_COLORS = {
        'SL':      ('7B2FBE', 'FFFFFF'),
        'CRITICO': ('FF4D4D', 'FFFFFF'),
        'ATENCAO': ('FF9800', 'FFFFFF'),
        'OK':      ('4CAF50', 'FFFFFF'),
        'NORMAL':  ('4CAF50', 'FFFFFF'),
        'ZERADO':  ('9E9E9E', 'FFFFFF'),
    }

    thin     = Side(style='thin',   color='CCCCCC')
    thin_d   = Side(style='thin',   color='0D0D1F')
    medium_d = Side(style='medium', color='0D0D1F')
    b_data   = Border(top=thin,   bottom=thin,     left=thin,   right=thin)
    b_hdr    = Border(top=thin_d, bottom=medium_d, left=thin_d, right=thin_d)

    col_widths = [10, 12, 12, 36, 18, 8, 14, 12, 12, 13, 14, 15, 14, 10, 19, 26, 26, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    NUM_COLS    = {7, 8, 9, 11, 12, 13}  # Qtde Log, Qtde Ant, Qtde Atual, Dias, V3M, VMes
    CENTER_COLS = {1, 2, 3, 6, 10, 11, 14, 15, 18}

    # Colunas que devem ser numéricas (1-based): Qtde Log, Qtde Atual, Dias, V3M, VMes
    COLS_INT   = {7, 8, 10, 11, 12}  # mesmas de NUM_COLS

    def to_num(v):
        """Converte string para int ou float; retorna None se vazio/inválido."""
        if v is None or str(v).strip() == '':
            return None
        try:
            f = float(str(v).replace(',', '.'))
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            return v   # mantém original se não for número

    for r_idx, row in enumerate(linhas, 1):
        ws.row_dimensions[r_idx].height = 22 if r_idx == 1 else 18
        is_hdr = (r_idx == 1)
        row_bg = ROW_ODD if r_idx % 2 == 0 else ROW_EVEN

        for c_idx, value in enumerate(row, 1):
            # Converte colunas numéricas para número real (evita bandeira verde do Excel)
            if not is_hdr and c_idx in COLS_INT:
                value = to_num(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if is_hdr:
                cell.font      = Font(bold=True, color=HEADER_FG, name='Segoe UI', size=10)
                cell.fill      = PatternFill('solid', fgColor=HEADER_BG)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border    = b_hdr
            else:
                cell.border = b_data
                if c_idx == 1:
                    sk = str(value or '').upper().strip()
                    if sk in STATUS_COLORS:
                        bg, fg = STATUS_COLORS[sk]
                        cell.fill = PatternFill('solid', fgColor=bg)
                        cell.font = Font(bold=True, color=fg, name='Segoe UI', size=9)
                    else:
                        cell.fill = PatternFill('solid', fgColor=row_bg)
                        cell.font = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c_idx in NUM_COLS:
                    cell.fill          = PatternFill('solid', fgColor=row_bg)
                    cell.font          = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0'
                elif c_idx in CENTER_COLS:
                    cell.fill      = PatternFill('solid', fgColor=row_bg)
                    cell.font      = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill      = PatternFill('solid', fgColor=row_bg)
                    cell.font      = Font(name='Segoe UI', size=9, color='222222')
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Atomo-ShelfLife.xlsx'
    )

@app.route('/api/pivot-clientes')
def pivot_clientes_novo():
    """
    Cruza FATURAMENTO x CARTEIRA com chave HIBRIDA (codigo do cliente + CNPJ).

    Prioridade de match:
      1) codigo E CNPJ iguais  -> match perfeito
      2) codigo igual          -> mesmo cadastro/base
      3) CNPJ igual            -> mesmo cliente com codigo diferente
    Em empate, ganha a carteira do vendedor filtrado.
    Venda sem nenhum match continua na lista (sem_carteira = true) e e
    atribuida ao vendedor gravado na propria nota, para o total bater
    sempre com a planilha.
    """
    try:
        vendedores = request.args.getlist('vendedor')
        produtos   = request.args.getlist('produtos')
        periodos   = request.args.getlist('periodo')
        anos       = request.args.getlist('ano')

        # ---------- filtro de periodo ----------
        periodo_conds  = []
        periodo_params = []
        if periodos:
            for per in periodos:
                try:
                    ano_p, mes_p = per.split('-')
                    periodo_conds.append('(f.ano = %s AND f.mes = %s)')
                    periodo_params += [int(ano_p), int(mes_p)]
                except:
                    pass
        elif anos:
            for a in anos:
                periodo_conds.append('f.ano = %s')
                periodo_params.append(int(a))

        if not periodo_conds and not periodos and not anos:
            from datetime import datetime as _dtnow
            periodo_conds.append('f.ano = %s')
            periodo_params.append(_dtnow.now().year)

        periodo_filter = ('AND (' + ' OR '.join(periodo_conds) + ')') if periodo_conds else ''

        # ---------- descobre a coluna de CNPJ em cada tabela ----------
        def _col(tabela, candidatos):
            try:
                rows = consultar(
                    "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
                    [tabela]
                )
            except Exception:
                return None
            existentes = {str(r['column_name']).lower() for r in rows}
            for c in candidatos:
                if c in existentes:
                    return c
            return None

        CAND   = ['cnpj_cpf', 'cnpj', 'cpf_cnpj', 'cnpj_cliente', 'documento', 'doc']
        c_cnpj = _col('carteira', CAND)
        f_cnpj = _col('faturamento', CAND)

        def nz(expr):
            # normaliza CNPJ/CPF: apenas digitos, sem zeros a esquerda
            return ("LTRIM(REGEXP_REPLACE(COALESCE(" + expr + "::TEXT, ''), '[^0-9]', '', 'g'), '0')")

        c_cnpj_sql = nz('c.' + c_cnpj) if c_cnpj else "''"
        f_cnpj_sql = nz('f.' + f_cnpj) if f_cnpj else "''"

        # ---------- codigos dos vendedores filtrados ----------
        cods = []
        if vendedores:
            vend_ph  = ','.join(['%s'] * len(vendedores))
            cod_rows = consultar(
                'SELECT DISTINCT cod_vendedor FROM vendedores WHERE nome IN (' + vend_ph + ') '
                'UNION '
                'SELECT DISTINCT cod_vendedor FROM faturamento WHERE vendedor IN (' + vend_ph + ') '
                'AND cod_vendedor IS NOT NULL',
                vendedores + vendedores
            )
            cods = [str(r['cod_vendedor']) for r in cod_rows if r.get('cod_vendedor') is not None]

        # ---------- CTEs (joins de conjunto, sem LATERAL, para nao travar) ----------
        # cart_cod : 1 linha por codigo de cliente (preferindo o vendedor filtrado)
        # cart_doc : 1 linha por CNPJ         (preferindo o vendedor filtrado)
        base_cte = (
            'WITH cart AS ('
            '  SELECT c.cod_cliente, c.cliente, c.endereco, c.cidade, c.uf,'
            '         c.cod_vendedor, ' + c_cnpj_sql + ' AS cnpj_n'
            '  FROM carteira c'
            '), cart_cod AS ('
            '  SELECT DISTINCT ON (cod_cliente) * FROM cart'
            '  ORDER BY cod_cliente, (cod_vendedor::TEXT = ANY(%s::TEXT[])) DESC, cnpj_n'
            '), cart_doc AS ('
            '  SELECT DISTINCT ON (cnpj_n) * FROM cart WHERE cnpj_n <> %s'
            '  ORDER BY cnpj_n, (cod_vendedor::TEXT = ANY(%s::TEXT[])) DESC, cod_cliente'
            '), fat AS ('
            '  SELECT f.cod_cliente, f.cliente, f.cidade, f.uf, f.bairro, f.ano, f.mes,'
            '         f.tipo_operacao, f.valor_nf, f.produto,'
            '         f.cod_vendedor AS fat_cod_vendedor,'
            '         ' + f_cnpj_sql + ' AS cnpj_n'
            '  FROM faturamento f'
            '  WHERE f.mes > 0 AND f.tipo_operacao IN (%s, %s)'
            '  ' + periodo_filter +
            '), m0 AS ('
            '  SELECT fat.*,'
            '         cc.cod_vendedor AS cc_vend, cc.cliente AS cc_cliente,'
            '         cc.cod_cliente  AS cc_cod,  cc.endereco AS cc_end, cc.cnpj_n AS cc_cnpj,'
            '         cd.cod_vendedor AS cd_vend, cd.cliente AS cd_cliente,'
            '         cd.cod_cliente  AS cd_cod,  cd.endereco AS cd_end'
            '  FROM fat'
            '  LEFT JOIN cart_cod cc ON cc.cod_cliente = fat.cod_cliente'
            '  LEFT JOIN cart_doc cd ON fat.cnpj_n <> %s AND cd.cnpj_n = fat.cnpj_n'
            '), m AS ('
            '  SELECT m0.*,'
            '    CASE WHEN cc_cod IS NOT NULL THEN cc_vend    ELSE cd_vend    END AS cart_cod_vendedor,'
            '    CASE WHEN cc_cod IS NOT NULL THEN cc_cliente ELSE cd_cliente END AS cart_cliente,'
            '    CASE WHEN cc_cod IS NOT NULL THEN cc_cod     ELSE cd_cod     END AS cart_cod_cliente,'
            '    CASE WHEN cc_cod IS NOT NULL THEN cc_end     ELSE cd_end     END AS cart_endereco'
            '  FROM m0'
            ')'
        )
        base_params = [cods, '', cods, 'Venda', 'Devolucao'] + periodo_params + ['']

        # ---------- filtro de vendedor ----------
        vend_filter = ''
        vend_params = []
        if cods:
            # vendas casadas com a carteira do vendedor + vendas orfas da nota dele
            vend_filter = (' AND (m.cart_cod_vendedor::TEXT = ANY(%s::TEXT[])'
                           ' OR (m.cart_cod_vendedor IS NULL AND m.fat_cod_vendedor::TEXT = ANY(%s::TEXT[])))')
            vend_params = [cods, cods]
        elif vendedores:
            vend_filter = ' AND v.nome IN (' + ','.join(['%s'] * len(vendedores)) + ')'
            vend_params = list(vendedores)

        select_cols = (
            ' SELECT COALESCE(m.cart_cliente, m.cliente) AS cliente,'
            ' m.cod_cliente, v.nome AS vendedor, m.ano, m.mes,'
            ' MAX(m.cidade) AS cidade, MAX(m.uf) AS uf, MAX(m.bairro) AS bairro,'
            ' MAX(m.cart_endereco) AS endereco,'
            ' BOOL_AND(m.cart_cod_vendedor IS NULL) AS sem_carteira,'
            ' ROUND(SUM(CASE WHEN m.tipo_operacao = %s THEN m.valor_nf ELSE 0 END)::NUMERIC,2) AS faturamento,'
            ' ROUND(SUM(CASE WHEN m.tipo_operacao = %s THEN m.valor_nf ELSE 0 END)::NUMERIC,2) AS devolucoes'
        )
        group_by = (' GROUP BY COALESCE(m.cart_cliente, m.cliente), m.cod_cliente, v.nome, m.ano, m.mes')

        conn   = get_conn()
        cursor = conn.cursor()

        if produtos:
            prod_ph = ','.join(['%s'] * len(produtos))
            sql = (
                base_cte + select_cols +
                ' FROM m'
                ' LEFT JOIN vendedores v ON v.cod_vendedor = m.cart_cod_vendedor'
                ' WHERE m.produto IN (' + prod_ph + ')'
                + vend_filter + group_by +
                ' ORDER BY 1, m.ano, m.mes'
            )
            cursor.execute(sql, base_params + ['Venda', 'Devolucao'] + list(produtos) + vend_params)

        else:
            # clientes da carteira do vendedor que nao compraram no periodo
            cart_zero   = ''
            zero_params = []
            if cods:
                cart_zero = (
                    ' UNION ALL'
                    ' SELECT c.cliente, c.cod_cliente, v.nome AS vendedor,'
                    ' NULL::INT AS ano, NULL::INT AS mes,'
                    ' c.cidade, c.uf, NULL::TEXT AS bairro, c.endereco, FALSE AS sem_carteira,'
                    ' 0::NUMERIC AS faturamento, 0::NUMERIC AS devolucoes'
                    ' FROM carteira c'
                    ' LEFT JOIN vendedores v ON v.cod_vendedor = c.cod_vendedor'
                    ' WHERE c.cod_vendedor::TEXT = ANY(%s::TEXT[])'
                    ' AND NOT EXISTS (SELECT 1 FROM m WHERE m.cart_cod_cliente = c.cod_cliente)'
                )
                zero_params = [cods]

            sql = (
                base_cte +
                ' SELECT * FROM (' + select_cols +
                ' FROM m'
                ' LEFT JOIN vendedores v ON v.cod_vendedor = m.cart_cod_vendedor'
                ' WHERE TRUE' + vend_filter + group_by +
                cart_zero +
                ' ) t ORDER BY cliente, ano, mes'
            )
            cursor.execute(sql, base_params + ['Venda', 'Devolucao'] + vend_params + zero_params)

        rows      = cursor.fetchall()
        cols      = [desc[0] for desc in cursor.description]
        resultado = [_serializar_row(dict(zip(cols, row))) for row in rows]
        cursor.close()
        conn.close()
        return jsonify(resultado)

    except Exception as e:
        import traceback as tb
        return jsonify({'erro': str(e), 'trace': tb.format_exc()}), 500
@app.route('/api/todos-clientes')
def todos_clientes():
    """Lista de clientes para busca local instantânea no frontend (igual /api/todos-produtos)."""
    key    = 'todos_clientes'
    cached = cache_get(key)
    if cached: return jsonify(cached)

    resultado = consultar("""
        SELECT DISTINCT cliente, cod_cliente
        FROM faturamento
        WHERE cliente IS NOT NULL AND cliente != ''
        ORDER BY cliente
    """)
    cache_set(key, resultado)
    return jsonify(resultado)


@app.route('/api/pivot-cliente-produto')
def pivot_cliente_produto():
    """
    Retorna o cruzamento Produto x Cliente x Período (ano/mês) com o valor comprado.
    Usado na página de 'Produtos por Cliente': permite filtrar vários clientes e vários
    períodos de uma vez, e ver quais produtos cada cliente comprou e quanto gastou.
    """
    try:
        clientes_cod = request.args.getlist('cod_cliente')
        periodos     = request.args.getlist('periodo')
        anos         = request.args.getlist('ano')

        if not clientes_cod:
            return jsonify({'erro': 'Selecione ao menos um cliente'}), 400

        cli_ph     = ','.join(['%s'] * len(clientes_cod))
        where_parts = ['f.cod_cliente IN (' + cli_ph + ')', 'f.mes > 0']
        params      = list(clientes_cod)

        # Monta filtro de período (mesma lógica do /api/pivot-clientes)
        periodo_conds  = []
        periodo_params = []
        if periodos:
            for per in periodos:
                try:
                    ano_p, mes_p = per.split('-')
                    periodo_conds.append('(f.ano = %s AND f.mes = %s)')
                    periodo_params += [int(ano_p), int(mes_p)]
                except:
                    pass
        elif anos:
            for a in anos:
                periodo_conds.append('f.ano = %s')
                periodo_params.append(int(a))

        if periodo_conds:
            where_parts.append('(' + ' OR '.join(periodo_conds) + ')')
            params += periodo_params

        sql = (
            'SELECT f.cliente, f.cod_cliente, f.produto, f.cod_produto, f.marca,'
            ' f.ano, f.mes,'
            ' ROUND(SUM(CASE WHEN f.tipo_operacao = %s THEN f.valor_nf ELSE 0 END)::NUMERIC,2) AS faturamento,'
            ' ROUND(SUM(CASE WHEN f.tipo_operacao = %s THEN f.valor_nf ELSE 0 END)::NUMERIC,2) AS devolucoes'
            ' FROM faturamento f'
            ' WHERE ' + ' AND '.join(where_parts) +
            ' GROUP BY f.cliente, f.cod_cliente, f.produto, f.cod_produto, f.marca, f.ano, f.mes'
            ' ORDER BY f.produto, f.cliente, f.ano, f.mes'
        )
        resultado = consultar(sql, ['Venda', 'Devolucao'] + params)
        return jsonify(resultado)

    except Exception as e:
        import traceback as tb
        return jsonify({'erro': str(e), 'trace': tb.format_exc()}), 500
def shelflife_historico():
    shelflife_id = request.args.get('shelflife_id')
    cod_produto  = request.args.get('cod_produto')
    where        = []; params = []

    if shelflife_id:
        where.append("shelflife_id = %s"); params.append(shelflife_id)
    if cod_produto:
        where.append("cod_produto = %s");  params.append(cod_produto)

    where_str = "WHERE " + " AND ".join(where) if where else ""
    resultado = consultar(f"""
        SELECT * FROM shelflife_historico {where_str}
        ORDER BY created_at DESC LIMIT 50
    """, params)
    return jsonify(resultado)

# Limpa cache (útil após atualizar dados)
@app.route('/api/cache/clear', methods=['GET', 'POST'])
def limpar_cache():
    cache_clear()
    return jsonify({"status": "cache limpo!"})

# Ping — mantém o servidor acordado
@app.route('/ping')
def ping():
    # Aquece o cache em background se estiver frio
    if 'filtros' not in _cache:
        _aquecer_cache()
    return jsonify({"status": "pong", "uptime": "ok"})


# ============================================================
#  VALTER — PESQUISA DE PRODUTO
#  Busca informações do produto nos sites da Alimentare
#  antes de recorrer à internet geral
# ============================================================
@app.route('/api/valter/pesquisar')
def valter_pesquisar():
    import requests as _req
    import urllib.parse

    produto     = request.args.get('produto', '').strip()
    if not produto:
        return jsonify({'erro': 'Produto não informado'}), 400

    SERPER_KEY  = os.environ.get('SERPER_API_KEY', '')
    resultados  = []

    SITES_ALIMENTARE = ['alimentareshop.com.br','alimentare.com.br','esdel.com.br','reforpan.com.br']

    def buscar_serper(query, num=5):
        try:
            r = _req.post(
                'https://google.serper.dev/search',
                headers={'X-API-KEY': SERPER_KEY, 'Content-Type': 'application/json'},
                json={'q': query, 'gl': 'br', 'hl': 'pt', 'num': num},
                timeout=10
            )
            if r.status_code == 200:
                return r.json()
        except Exception:
            pass
        return {}

    if SERPER_KEY:
        # 1. Busca nos sites da Alimentare primeiro
        query_sites = f'{produto} site:alimentareshop.com.br OR site:alimentare.com.br OR site:esdel.com.br OR site:reforpan.com.br'
        data = buscar_serper(query_sites)
        for item in data.get('organic', [])[:4]:
            fonte = next((s for s in SITES_ALIMENTARE if s in item.get('link','')), 'Alimentare')
            resultados.append({
                'fonte':     fonte,
                'url':       item.get('link',''),
                'titulo':    item.get('title', produto),
                'descricao': item.get('snippet',''),
            })

        # 2. Se não achou, busca geral na internet
        if not resultados:
            data = buscar_serper(f'{produto} produto alimentício ingredientes calorias código barras')
            for item in data.get('organic', [])[:4]:
                resultados.append({
                    'fonte':     'Internet',
                    'url':       item.get('link',''),
                    'titulo':    item.get('title', produto),
                    'descricao': item.get('snippet',''),
                })

        # 3. Inclui knowledge graph se disponível
        kg = data.get('knowledgeGraph', {})
        if kg.get('description'):
            resultados.insert(0, {
                'fonte':     'Google',
                'url':       kg.get('website',''),
                'titulo':    kg.get('title', produto),
                'descricao': kg.get('description',''),
            })

    else:
        resultados.append({
            'fonte':     'Sistema',
            'url':       '',
            'titulo':    produto,
            'descricao': 'Configure SERPER_API_KEY no Railway para habilitar pesquisa.',
        })

    return jsonify({
        'produto':           produto,
        'resultados':        resultados,
        'total':             len(resultados),
        'fontes_consultadas': SITES_ALIMENTARE + ['Internet'],
    })

# ============================================================
#  VALTER — CONTEXTO COMPLETO DO SISTEMA
# ============================================================
@app.route('/api/valter/contexto')
def valter_contexto():
    try:
        from datetime import datetime as _dt
        ano  = _dt.now().year
        mes  = _dt.now().month
        conn = get_conn(); cur = conn.cursor()

        cur.execute("""
            SELECT
                ROUND(SUM(CASE WHEN tipo_operacao='Venda'     THEN valor_nf ELSE 0 END)::NUMERIC,2),
                ROUND(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END)::NUMERIC,2),
                COUNT(DISTINCT cod_cliente), COUNT(DISTINCT vendedor)
            FROM faturamento WHERE ano=%s""", [ano])
        k = cur.fetchone() or [0,0,0,0]

        cur.execute("""
            SELECT
                ROUND(SUM(CASE WHEN tipo_operacao='Venda'     THEN valor_nf ELSE 0 END)::NUMERIC,2),
                ROUND(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END)::NUMERIC,2)
            FROM faturamento WHERE ano=%s AND mes=%s""", [ano, mes])
        km = cur.fetchone() or [0,0]

        cur.execute("""
            SELECT vendedor, cod_vendedor,
                ROUND(SUM(CASE WHEN tipo_operacao='Venda'     THEN valor_nf ELSE 0 END)::NUMERIC,2),
                ROUND(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END)::NUMERIC,2),
                COUNT(DISTINCT cod_cliente)
            FROM faturamento WHERE ano=%s AND vendedor IS NOT NULL
            GROUP BY vendedor, cod_vendedor ORDER BY 3 DESC LIMIT 10""", [ano])
        vends = cur.fetchall()

        cur.execute("""
            SELECT marca,
                ROUND(SUM(CASE WHEN tipo_operacao='Venda' THEN valor_nf ELSE 0 END)::NUMERIC,2),
                COUNT(DISTINCT cod_cliente)
            FROM faturamento WHERE ano=%s AND marca IS NOT NULL
            GROUP BY marca ORDER BY 2 DESC LIMIT 10""", [ano])
        marcas = cur.fetchall()

        cur.execute("""
            SELECT produto, cod_produto,
                ROUND(SUM(CASE WHEN tipo_operacao='Venda' THEN valor_nf ELSE 0 END)::NUMERIC,2),
                SUM(CASE WHEN tipo_operacao='Venda' THEN quantidade ELSE 0 END)
            FROM faturamento WHERE ano=%s AND produto IS NOT NULL
            GROUP BY produto, cod_produto ORDER BY 3 DESC LIMIT 10""", [ano])
        prods = cur.fetchall()

        cur.execute("""
            SELECT mes,
                ROUND(SUM(CASE WHEN tipo_operacao='Venda'     THEN valor_nf ELSE 0 END)::NUMERIC,2),
                ROUND(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END)::NUMERIC,2)
            FROM faturamento WHERE ano=%s GROUP BY mes ORDER BY mes""", [ano])
        meses_fat = cur.fetchall()

        try:
            cur.execute("""
                SELECT COUNT(*),
                    SUM(CASE WHEN dias_vencimento<=30 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN is_sl=TRUE THEN 1 ELSE 0 END),
                    SUM(CASE WHEN (venda_mes IS NULL OR venda_mes=0) AND quantidade_log>0 THEN 1 ELSE 0 END)
                FROM shelflife WHERE semana=(SELECT MAX(semana) FROM shelflife)""")
            sl = cur.fetchone() or [0,0,0,0]
        except Exception:
            sl = [0,0,0,0]

        cur.execute("SELECT COUNT(*), COUNT(DISTINCT cod_vendedor) FROM carteira")
        cart = cur.fetchone() or [0,0]
        cur.close(); conn.close()

        meses_nome = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
                      7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

        ctx = (
            f"SISTEMA ATOMO — DADOS COMPLETOS ({ano})\n"
            f"FATURAMENTO {ano}: Vendas R$ {k[0]:,.2f} | Dev R$ {k[1]:,.2f} | Liquido R$ {(k[0]-k[1]):,.2f} | Clientes {k[2]:,} | Vendedores {k[3]}\n"
            f"MES ATUAL ({meses_nome.get(mes,'')}/{ano}): Vendas R$ {km[0]:,.2f} | Dev R$ {km[1]:,.2f}\n\n"
            "FATURAMENTO MENSAL:\n" +
            "\n".join(f"  {meses_nome.get(r[0],r[0])}: R$ {r[1]:,.2f} | Dev R$ {r[2]:,.2f}" for r in meses_fat) +
            "\n\nTOP 10 VENDEDORES (cod|nome|fat|dev|clientes):\n" +
            "\n".join(f"  {i+1}. Cod {r[1]} {r[0]} | R$ {r[2]:,.2f} | Dev R$ {r[3]:,.2f} | {r[4]} clientes" for i,r in enumerate(vends)) +
            "\n\nTOP 10 MARCAS:\n" +
            "\n".join(f"  {i+1}. {r[0]} | R$ {r[1]:,.2f} | {r[2]} clientes" for i,r in enumerate(marcas)) +
            "\n\nTOP 10 PRODUTOS:\n" +
            "\n".join(f"  {i+1}. {r[0]} (cod:{r[1]}) | R$ {r[2]:,.2f} | {int(r[3] or 0):,} un" for i,r in enumerate(prods)) +
            f"\n\nSHELF LIFE: Total {sl[0]} | Criticos {sl[1]} | Em SL {sl[2]} | Sem giro {sl[3]}"
            f"\nCARTEIRA: {cart[0]:,} clientes | {cart[1]} vendedores"
        )
        return jsonify({'contexto': ctx, 'ok': True})
    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'trace': traceback.format_exc()}), 500

# ============================================================
#  VALTER — ALERTAS PROATIVOS
# ============================================================
@app.route('/api/valter/alertas')
def valter_alertas():
    try:
        conn = get_conn(); cur = conn.cursor()
        alertas = []
        try:
            cur.execute("SELECT COUNT(*) FROM shelflife WHERE semana=(SELECT MAX(semana) FROM shelflife) AND dias_vencimento<=7 AND dias_vencimento>=0 AND is_sl=TRUE")
            n = cur.fetchone()[0]
            if n > 0:
                alertas.append({'icone':'🚨','texto':f'{n} produto{"s" if n>1 else ""} SL vencem em 7 dias','pergunta':f'Liste os {n} produtos SL que vencem em 7 dias.'})
            cur.execute("SELECT COUNT(*) FROM shelflife WHERE semana=(SELECT MAX(semana) FROM shelflife) AND (status_logistica='CRITICO' OR dias_vencimento<=30) AND (acao IS NULL OR acao='' OR acao='Lembrete')")
            n2 = cur.fetchone()[0]
            if n2 > 0:
                alertas.append({'icone':'⚠️','texto':f'{n2} produto{"s" if n2>1 else ""} crítico{"s" if n2>1 else ""} sem ação','pergunta':f'Quais os {n2} produtos críticos sem ação?'})
        except Exception:
            pass
        cur.close(); conn.close()
        return jsonify({'alertas': alertas})
    except Exception as e:
        return jsonify({'alertas': [], 'erro': str(e)})

# ============================================================
#  VALTER — PROXY CHAT
# ============================================================
@app.route('/api/valter/chat', methods=['POST'])
def valter_chat():
    import requests as _req
    try:
        data    = request.get_json(force=True)
        api_key = os.environ.get('GROQ_API_KEY', '')
        if not api_key:
            return jsonify({'erro': 'GROQ_API_KEY nao configurada'}), 500
        msgs = []
        if data.get('system'):
            msgs.append({'role':'system','content':data['system']})
        msgs += data.get('messages', [])
        resp = _req.post(
            'https://api.groq.com/openai/v1/chat/completions',
            headers={'Authorization':f'Bearer {api_key}','Content-Type':'application/json'},
            json={'model':'llama-3.3-70b-versatile','max_tokens':1000,'messages':msgs,'temperature':0.7},
            timeout=60
        )
        d = resp.json()
        texto = d['choices'][0]['message']['content'] if d.get('choices') else ''
        return jsonify({'content':[{'type':'text','text':texto}]}), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 500


# ============================================================
#  CLIENTES EM RISCO
# ============================================================
@app.route('/api/clientes-em-risco')
def clientes_em_risco():
    try:
        dias  = int(request.args.get('dias', 60))
        limit = int(request.args.get('limite', 200))
        conn  = get_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT f.cod_cliente, f.cliente, f.vendedor, f.cod_vendedor,
                MAX(f.data_movimento)::TEXT AS ultima_compra,
                (CURRENT_DATE - MAX(f.data_movimento::DATE))::INT AS dias_sem_compra,
                ROUND(SUM(CASE WHEN f.tipo_operacao='Venda' THEN f.valor_nf ELSE 0 END)::NUMERIC,2) AS fat_total,
                COUNT(DISTINCT f.data_movimento) AS num_pedidos
            FROM faturamento f
            WHERE f.tipo_operacao='Venda' AND f.cliente IS NOT NULL
            GROUP BY f.cod_cliente, f.cliente, f.vendedor, f.cod_vendedor
            HAVING (CURRENT_DATE - MAX(f.data_movimento::DATE))::INT >= %s
            ORDER BY dias_sem_compra DESC
            LIMIT %s
        """, [dias, limit])
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            row = dict(zip(cols, r))
            row['dias_sem_compra'] = int(row['dias_sem_compra'] or 0)
            row['fat_total']       = float(row['fat_total'] or 0)
            row['num_pedidos']     = int(row['num_pedidos'] or 0)
            rows.append(row)
        cur.close(); conn.close()
        return jsonify({'clientes': rows, 'total': len(rows), 'dias_corte': dias})
    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'trace': traceback.format_exc()}), 500

# ============================================================
#  COMPARAÇÃO DE PERÍODOS
# ============================================================
@app.route('/api/comparar-periodos')
def comparar_periodos():
    try:
        ano_a = request.args.get('ano_a')
        mes_a = request.args.get('mes_a')
        ano_b = request.args.get('ano_b')
        mes_b = request.args.get('mes_b')

        def buscar(ano, mes):
            where = ["tipo_operacao IN ('Venda','Devolucao')"]
            params = []
            if ano: where.append('ano = %s'); params.append(int(ano))
            if mes: where.append('mes = %s'); params.append(int(mes))
            sql = """SELECT
                ROUND(SUM(CASE WHEN tipo_operacao='Venda'     THEN valor_nf ELSE 0 END)::NUMERIC,2),
                ROUND(SUM(CASE WHEN tipo_operacao='Devolucao' THEN valor_nf ELSE 0 END)::NUMERIC,2),
                COUNT(DISTINCT cod_cliente),
                COUNT(DISTINCT num_nf),
                ROUND(AVG(CASE WHEN tipo_operacao='Venda' THEN valor_nf END)::NUMERIC,2)
                FROM faturamento WHERE """ + ' AND '.join(where)
            conn = get_conn(); cur = conn.cursor()
            cur.execute(sql, params)
            r = cur.fetchone(); cur.close(); conn.close()
            fat = float(r[0] or 0); dev = float(r[1] or 0)
            return {'faturamento': fat, 'devolucoes': dev, 'liquido': fat + dev,
                    'clientes': int(r[2] or 0), 'pedidos': int(r[3] or 0),
                    'ticket_medio': float(r[4] or 0)}

        pA = buscar(ano_a, mes_a)
        pB = buscar(ano_b, mes_b)

        def var(a, b):
            if not b: return None
            return round((a - b) / abs(b) * 100, 1)

        return jsonify({
            'periodo_a': pA, 'periodo_b': pB,
            'variacao': {k: var(pA[k], pB[k]) for k in pA}
        })
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

# ============================================================
#  HISTÓRICO DE COMPRAS DO CLIENTE
# ============================================================
# ============================================================
#  HISTÓRICO DE COMPRAS POR VENDEDOR / CLIENTE
# ============================================================
@app.route('/api/clientes/historico')
def cliente_historico():
    try:
        cod_cliente = request.args.get('cod_cliente', '').strip()
        cliente_nome = request.args.get('cliente', '').strip()
        vendedor_nome = request.args.get('vendedor', '').strip()

        where = []
        params = []

        if cod_cliente:
            where.append("cod_cliente = %s")
            params.append(cod_cliente)
        elif cliente_nome:
            where.append("LOWER(cliente) LIKE LOWER(%s)")
            params.append(f"%{cliente_nome}%")
        elif vendedor_nome:
            where.append("LOWER(vendedor) LIKE LOWER(%s)")
            params.append(f"%{vendedor_nome}%")

        where_str = "WHERE " + " AND ".join(where) if where else ""

        resultado = consultar(f"""
            SELECT 
                ano, 
                mes, 
                data_movimento, 
                num_nf, 
                tipo_operacao, 
                produto, 
                cod_produto, 
                marca, 
                quantidade, 
                ROUND(CAST(valor_nf AS NUMERIC), 2) AS valor_nf,
                vendedor,
                unidade,
                cod_cliente,
                cliente
            FROM faturamento 
            {where_str}
            ORDER BY data_movimento DESC, ano DESC, mes DESC
        """, params)

        return jsonify(resultado)

    except Exception as e:
        import traceback as tb
        return jsonify({'erro': str(e), 'trace': tb.format_exc()}), 500

@app.route('/api/compras/upload', methods=['POST'])
def compras_upload():
    """Recebe os itens da planilha (JSON) e grava no Railway."""
    conn = None
    try:
        data = request.get_json(force=True)
        items = data.get('items', [])
        if not items:
            return jsonify({'success': False, 'message': 'Nenhum item enviado'}), 400

        conn = get_conn()
        cursor = conn.cursor()

        # Último upload (para comparação)
        cursor.execute("SELECT id FROM purchases_uploads ORDER BY id DESC LIMIT 1")
        row = cursor.fetchone()
        ultimo_upload_id = row[0] if row else None

      # Novo lote — busca o primeiro usuário disponível no banco
        cursor.execute("SELECT id FROM users LIMIT 1")
        user_row = cursor.fetchone()
        user_id = user_row[0] if user_row else None
        if not user_id:
            return jsonify({'success': False, 'message': 'Nenhum usuário cadastrado no banco'}), 500

        # DEPOIS
        cursor.execute("""
            INSERT INTO purchases_uploads (userid, filename, totalitems)
            VALUES (%s, 'Compras_Diarias.xlsx', %s) RETURNING id
        """, [user_id, len(items)])
        novo_upload_id = cursor.fetchone()[0]

        # Itens
        for item in items:
            cursor.execute("""
                INSERT INTO purchase_items
                    (uploadid, codigoproduto, descricaoproduto, quantidademediavenda, marca, carimbodatahora)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, [
                novo_upload_id,
                str(item.get('codigoProduto', '')),
                str(item.get('nomeProduto', '')),
                str(item.get('quantidade', '0')),
                str(item.get('total', '0')),
                str(item.get('dataHora', '')),
            ])

        # KPIs comparativos
        codigos_atuais = [str(i.get('codigoProduto', '')).strip()
                          for i in items if i.get('codigoProduto')]

        existiam_na_anterior = 0
        if ultimo_upload_id:
            cursor.execute(
                "SELECT DISTINCT codigoproduto FROM purchase_items WHERE uploadid = %s",
                [ultimo_upload_id]
            )
            anteriores = {str(r[0]).strip() for r in cursor.fetchall()}
            existiam_na_anterior = sum(1 for c in codigos_atuais if c in anteriores)

        novos_produtos = len(codigos_atuais) - existiam_na_anterior

        conn.commit()
        cursor.close(); conn.close()
        cache_clear()

        return jsonify({
            'success': True,
            'uploadId': novo_upload_id,
            'analise': {
                'totalPlanilha': len(items),
                'existiamNaAnterior': existiam_na_anterior,
                'novosProdutos': novos_produtos
            }
        })
    except Exception as e:
        if conn:
            try: conn.rollback()
            except Exception: pass
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/compras/snapshots', methods=['GET'])
def compras_snapshots():
    """Lista todas as planilhas enviadas (mais recentes primeiro)."""
    try:
        conn = get_conn()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute("""
            SELECT id, fileName, totalItems,
                   to_char(uploadDate, 'DD/MM/YYYY HH24:MI') AS uploaddate
            FROM purchases_uploads
            ORDER BY id DESC
        """)
        rows = cursor.fetchall()
        cursor.close(); conn.close()
        return jsonify({'snapshots': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'snapshots': [], 'error': str(e)}), 500


@app.route('/api/compras/listar', methods=['GET'])
def compras_listar():
    """Retorna os itens de um upload + códigos da planilha anterior (para marcar 'NOVO')."""
    try:
        upload_id = int(request.args.get('uploadId'))
        conn = get_conn()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        cursor.execute("""
            SELECT codigoProduto, descricaoProduto, quantidadeMediaVenda, marca
            FROM purchase_items
            WHERE uploadId = %s
            ORDER BY id
        """, [upload_id])
        items = [dict(r) for r in cursor.fetchall()]

        cursor.execute("""
            SELECT id FROM purchases_uploads
            WHERE id < %s
            ORDER BY id DESC LIMIT 1
        """, [upload_id])
        anterior = cursor.fetchone()

        codigos_anteriores = []
        if anterior:
            cursor.execute(
                "SELECT DISTINCT codigoProduto FROM purchase_items WHERE uploadId = %s",
                [anterior['id']]
            )
            codigos_anteriores = [str(r['codigoproduto']).strip()
                                  for r in cursor.fetchall()]

        cursor.close(); conn.close()
        return jsonify({'items': items, 'codigosAnteriores': codigos_anteriores})
    except Exception as e:
        return jsonify({'items': [], 'codigosAnteriores': [], 'error': str(e)}), 500


@app.route('/api/compras/upload/<int:upload_id>', methods=['DELETE'])
def compras_deletar(upload_id):
    """Exclui uma planilha (cascata apaga os itens)."""
    try:
        conn = get_conn()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM purchases_uploads WHERE id = %s", [upload_id])
        conn.commit()
        cursor.close(); conn.close()
        cache_clear()
        return jsonify({'success': True, 'message': 'Planilha excluída com sucesso'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/compras/exportar', methods=['GET'])
def compras_exportar():
    """Exporta os itens de um upload em .xlsx."""
    try:
        import openpyxl
        upload_id = int(request.args.get('uploadId'))
        conn = get_conn()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cursor.execute("""
            SELECT codigoProduto, descricaoProduto, quantidadeMediaVenda, marca
            FROM purchase_items
            WHERE uploadId = %s
            ORDER BY id
        """, [upload_id])
        rows = cursor.fetchall()
        cursor.close(); conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Compras'
        ws.append(['Código', 'Produto', 'Quantidade', 'Fabricante'])
        for r in rows:
            ws.append([r['codigoproduto'], r['descricaoproduto'],
                       r['quantidademediavenda'], r['marca']])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f'compras_{upload_id}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500